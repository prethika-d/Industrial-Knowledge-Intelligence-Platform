"""
Document Processing Service
Reads uploaded PDF, DOCX, TXT and XLSX files,
extracts their text, stores it in the database,
and marks the document as indexed.
"""

import mimetypes
import os

from PyPDF2 import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

from .models import Document, FileType, ProcessingStatus


EXTENSION_TO_FILE_TYPE = {
    "pdf": FileType.PDF,
    "docx": FileType.DOCX,
    "doc": FileType.DOC,
    "txt": FileType.TXT,
    "xlsx": FileType.XLSX,
}


def infer_file_type(filename):
    ext = filename.rsplit(".", 1)[-1].lower()
    return EXTENSION_TO_FILE_TYPE.get(ext, FileType.TXT)

def infer_category(filename):
    filename = filename.lower()

    if "manual" in filename:
        return "manual"

    elif "sop" in filename:
        return "sop"

    elif "inspection" in filename:
        return "inspection_report"

    elif "maintenance" in filename:
        return "maintenance"

    elif "safety" in filename:
        return "safety"

    return "other"
def guess_mimetype(filename):
    mime, _ = mimetypes.guess_type(filename)
    return mime or "application/octet-stream"


def extract_pdf(path):
    text = ""

    reader = PdfReader(path)

    for page in reader.pages:
        page_text = page.extract_text()

        if page_text:
            text += page_text + "\n"

    return text


def extract_docx(path):
    doc = DocxDocument(path)

    return "\n".join(
        paragraph.text
        for paragraph in doc.paragraphs
    )


def extract_txt(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def extract_xlsx(path):
    workbook = load_workbook(path)

    text = ""

    for sheet in workbook.worksheets:

        text += f"\nSheet: {sheet.title}\n"

        for row in sheet.iter_rows(values_only=True):

            row_text = " ".join(
                str(cell)
                for cell in row
                if cell is not None
            )

            text += row_text + "\n"

    return text


def generate_summary(text):

    if not text:
        return "No text extracted."

    lines = text.splitlines()

    summary = []

    for line in lines:

        line = line.strip()

        if line:

            summary.append(line)

        if len(summary) == 8:
            break

    return "\n".join(summary)


def process_document(document):

    document.processing_status = ProcessingStatus.PROCESSING
    document.save()

    path = document.file.path

    print("FILE PATH:", path)

    extension = os.path.splitext(path)[1].lower()

    print("EXTENSION:", extension)

    extracted = ""

    try:

        if extension == ".pdf":
            print("Reading PDF...")
            extracted = extract_pdf(path)

        elif extension == ".docx":
            print("Reading DOCX...")
            extracted = extract_docx(path)

        elif extension == ".txt":
            print("Reading TXT...")
            extracted = extract_txt(path)

        elif extension == ".xlsx":
            print("Reading XLSX...")
            extracted = extract_xlsx(path)

        else:
            extracted = "Unsupported document type."

        print("TEXT LENGTH:", len(extracted))
        print(extracted[:500])

        document.extracted_text = extracted
        document.summary = generate_summary(extracted)
        document.processing_status = ProcessingStatus.INDEXED
        document.save()

    except Exception as e:
        print("ERROR:", e)

        document.processing_status = ProcessingStatus.FAILED
        document.summary = str(e)
        document.save()

    return document